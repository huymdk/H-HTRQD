# File: app.py
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
import db_manager
import ahp_core
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import io
import json

app = Flask(__name__)
app.secret_key = 'my_ahp_electronics_secret_key_v12_final_final_final_!@#$' # Cập nhật key

NUM_CRITERIA = 4
CRITERIA_NAMES = ["Hiệu suất", "Đánh giá", "Thiết kế", "Giá cả"]

# --- Helper Functions ---
def format_float(value, precision=4):
    return f"{value:.{precision}f}" if isinstance(value, (int, float)) else str(value)

def _ensure_python_types_for_session(data_dict):
    if data_dict is None: return None
    processed_dict = {}
    for key, value in data_dict.items():
        if isinstance(value, np.ndarray): processed_dict[key] = value.tolist()
        elif isinstance(value, np.integer): processed_dict[key] = int(value)
        elif isinstance(value, np.floating): processed_dict[key] = float(value)
        elif isinstance(value, np.bool_): processed_dict[key] = bool(value)
        elif isinstance(value, bool): processed_dict[key] = value
        elif isinstance(value, list):
            new_list = []
            for item in value:
                if isinstance(item, dict): new_list.append(_ensure_python_types_for_session(item))
                elif isinstance(item, np.integer): new_list.append(int(item))
                elif isinstance(item, np.floating): new_list.append(float(item))
                elif isinstance(item, np.bool_): new_list.append(bool(item))
                elif isinstance(item, bool): new_list.append(item)
                else: new_list.append(item)
            processed_dict[key] = new_list
        elif isinstance(value, dict): processed_dict[key] = _ensure_python_types_for_session(value)
        else: processed_dict[key] = value
    return processed_dict

def _calculate_criteria_logic(form_data):
    num_crit = NUM_CRITERIA; crit_comparison_values = []
    for i in range(num_crit):
        for j in range(i + 1, num_crit):
            val_str = form_data.get(f'compare_{i}_{j}');
            if not val_str: raise ValueError(f"Thiếu giá trị so sánh cho '{CRITERIA_NAMES[i]}' và '{CRITERIA_NAMES[j]}'.")
            val = int(val_str);
            if not (1 <= val <= 9): raise ValueError("Giá trị so sánh tiêu chí phải từ 1 đến 9.")
            crit_comparison_values.append(val)
    crit_matrix_np = ahp_core.create_comparison_matrix(num_crit); ahp_core.fill_comparison_matrix_from_list(crit_matrix_np, crit_comparison_values)
    crit_weights_np = ahp_core.calculate_priority_vector(crit_matrix_np); lambda_max_c, ci_c, cr_c = ahp_core.check_consistency(crit_matrix_np, crit_weights_np)
    col_sums = np.sum(crit_matrix_np, axis=0); normalized_crit_matrix_by_col_np = crit_matrix_np / col_sums if np.all(col_sums != 0) else np.zeros_like(crit_matrix_np)
    return _ensure_python_types_for_session({
        'criteria_names': CRITERIA_NAMES, 'comparison_matrix': crit_matrix_np.tolist(), 'weights': crit_weights_np.tolist(),
        'lambda_max': lambda_max_c, 'ci': ci_c, 'cr': cr_c, # Key 'cr' cho CR của tiêu chí
        'normalized_matrix_by_col': normalized_crit_matrix_by_col_np.tolist(), 'is_consistent': (cr_c <= 0.1)})

def _calculate_full_ahp_logic(form_data, criteria_results_dict):
    crit_weights_list = criteria_results_dict['weights']; crit_weights_np = np.array(crit_weights_list)
    num_crit = NUM_CRITERIA; num_alternatives = int(form_data.get('num_alternatives', 3)); alt_names = [form_data.get(f'alt_name_{k}', f'PA {k+1}') for k in range(num_alternatives)]
    raw_scores_data_np = np.zeros((num_alternatives, num_crit)); criteria_types_list = [form_data.get(f'crit_type_{j}', 'benefit') for j in range(num_crit)]
    for j in range(num_crit):
        for i in range(num_alternatives):
            score_str = form_data.get(f'score_{i}_{j}');
            if not score_str: raise ValueError(f"Thiếu điểm cho '{alt_names[i]}' theo '{CRITERIA_NAMES[j]}'.")
            raw_scores_data_np[i, j] = float(score_str)
    normalized_alt_scores_np = ahp_core.normalize_alternative_scores(raw_scores_data_np, criteria_types_list)
    overall_scores_np = ahp_core.calculate_overall_scores(normalized_alt_scores_np, crit_weights_np)
    ranked_alternatives_info_list = []
    for i in range(num_alternatives):
        ranked_alternatives_info_list.append({'name': alt_names[i], 'raw_scores': raw_scores_data_np[i,:].tolist(), 'normalized_scores': normalized_alt_scores_np[i,:].tolist(), 'overall_score': overall_scores_np[i]})
    ranked_alternatives_info_list.sort(key=lambda x: x['overall_score'], reverse=True)
    for rank, alt_info in enumerate(ranked_alternatives_info_list): alt_info['rank'] = rank + 1
    full_results = {**criteria_results_dict, 'ranked_alternatives': ranked_alternatives_info_list, 'num_alternatives_processed': num_alternatives, 'criteria_types': criteria_types_list,
                    'chart_crit_names': criteria_results_dict['criteria_names'], 'chart_crit_weights': criteria_results_dict['weights'],
                    'chart_alt_names': [alt['name'] for alt in ranked_alternatives_info_list], 'chart_alt_scores': [alt['overall_score'] for alt in ranked_alternatives_info_list]}
    return _ensure_python_types_for_session(full_results)

def store_ahp_results_to_db(decision_name, description, results_data_dict):
    crit_comparison_matrix_json = json.dumps(results_data_dict['comparison_matrix']) if 'comparison_matrix' in results_data_dict else None
    # Sử dụng key 'cr' khi lấy CR từ results_data_dict
    decision_id = db_manager.add_decision(name=decision_name, description=description, criteria_cr_value=results_data_dict.get('cr'), criteria_comparison_matrix_json=crit_comparison_matrix_json)
    if not decision_id: flash("Lỗi CSDL: Không thể lưu quyết định chính.", "error"); return None
    criteria_db_objects_map = {}
    for i, crit_name in enumerate(results_data_dict.get('criteria_names', [])):
        weight = results_data_dict.get('weights', [])[i] if results_data_dict.get('weights') and i < len(results_data_dict.get('weights', [])) else None
        crit_id = db_manager.add_criterion(decision_id, crit_name, weight)
        if crit_id: criteria_db_objects_map[crit_name] = crit_id
        else: flash(f"Lỗi CSDL khi lưu tiêu chí '{crit_name}'.", "error")
    for alt_data in results_data_dict.get('ranked_alternatives', []):
        alt_id = db_manager.add_alternative(decision_id, alt_data['name'])
        if alt_id:
            normalized_scores_for_db = {}
            for j, crit_name_for_score in enumerate(results_data_dict.get('criteria_names', [])):
                criterion_id_for_score = criteria_db_objects_map.get(crit_name_for_score)
                if criterion_id_for_score:
                    raw_score_val = alt_data.get('raw_scores', [])[j] if alt_data.get('raw_scores') and j < len(alt_data.get('raw_scores',[])) else None
                    norm_score_val = alt_data.get('normalized_scores', [])[j] if alt_data.get('normalized_scores') and j < len(alt_data.get('normalized_scores',[])) else None
                    if raw_score_val is not None: db_manager.save_alternative_raw_score(alt_id, criterion_id_for_score, raw_score_val)
                    if norm_score_val is not None: normalized_scores_for_db[criterion_id_for_score] = norm_score_val
                else: flash(f"Không tìm thấy ID cho tiêu chí '{crit_name_for_score}' khi lưu điểm PA.", "warning")
            db_manager.update_alternative_details(alt_id, normalized_scores_for_db, alt_data.get('overall_score',0), alt_data.get('rank',0))
        else: flash(f"Lỗi CSDL khi lưu phương án '{alt_data['name']}'.", "error")
    return decision_id

@app.context_processor
def utility_processor():
    return dict(_datetime=datetime, format_float=format_float)

@app.route('/', methods=['GET', 'POST'])
def index():
    criteria_results = session.get('criteria_results', None)
    ahp_final_results = None
    submitted_data = {}
    if request.method == 'POST':
        submitted_data = request.form.to_dict(); session['last_form_data'] = submitted_data
        action = request.form.get('action')
        try:
            if action == 'calculate_criteria':
                criteria_results = _calculate_criteria_logic(request.form); session['criteria_results'] = criteria_results
                flash("Đã tính toán và kiểm tra nhất quán cho tiêu chí.", "info")
            elif action == 'calculate_full_ahp':
                criteria_data_for_full_calc = session.get('criteria_results')
                if not criteria_data_for_full_calc or 'weights' not in criteria_data_for_full_calc:
                    flash("Lỗi: Vui lòng 'Tính Trọng Số Tiêu Chí' trước.", "error")
                else:
                    ahp_final_results = _calculate_full_ahp_logic(request.form, criteria_data_for_full_calc)
                    decision_name_for_db = f"Đồ điện tử (Web) - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                    decision_description_for_db = "Kết quả AHP từ trang web."
                    saved_decision_id = store_ahp_results_to_db(decision_name_for_db, decision_description_for_db, ahp_final_results)
                    if saved_decision_id: flash(f"Đã tính toán và lưu kết quả vào lịch sử (ID: {saved_decision_id}).", "success")
                    else: flash("Tính toán AHP thành công, nhưng có lỗi khi lưu vào lịch sử.", "warning")
                    session.pop('criteria_results', None); session.pop('last_form_data', None)
            else: flash("Hành động không hợp lệ.", "error")
        except ValueError as e: flash(f"Lỗi nhập liệu: {e}", "error"); criteria_results = session.get('criteria_results', None)
        except Exception as e_gen: flash(f"Lỗi không xác định: {e_gen}", "error"); criteria_results = session.get('criteria_results', None)
    else: # GET request
        if 'last_form_data' in session and request.args.get('new_session') != 'true':
            submitted_data = session.get('last_form_data', {}); criteria_results = session.get('criteria_results', None)
        else: session.pop('criteria_results', None); session.pop('last_form_data', None); criteria_results = None; submitted_data = {}
    return render_template('index.html', criteria_names=CRITERIA_NAMES, num_criteria=NUM_CRITERIA,
                           criteria_results=criteria_results, ahp_final_results=ahp_final_results, submitted_data=submitted_data)

@app.route('/history')
def history():
    decisions_history = db_manager.get_all_decisions()
    return render_template('history.html', decisions=decisions_history)

@app.route('/decision_results/<int:decision_id>')
def decision_results_page(decision_id):
    decision_info = db_manager.get_decision_by_id(decision_id)
    if not decision_info: flash(f"Không tìm thấy quyết định với ID {decision_id}.","error"); return redirect(url_for('history'))
    criteria_from_db = db_manager.get_criteria_for_decision(decision_id)
    alternatives_from_db = db_manager.get_alternatives_with_scores_for_decision(decision_id)
    crit_comp_matrix_list_from_db = None
    if decision_info.get('criteria_comparison_json'):
        try: crit_comp_matrix_list_from_db = json.loads(decision_info['criteria_comparison_json'])
        except (json.JSONDecodeError, TypeError): flash("Lỗi đọc ma trận JSON từ lịch sử.","warning")

    processed_chart_crit_names = []
    processed_chart_crit_weights = []
    if criteria_from_db:
        temp_crit_names = [c.get('criterion_name', 'N/A') for c in criteria_from_db]
        temp_crit_weights = [c.get('weight') for c in criteria_from_db]
        for i, w in enumerate(temp_crit_weights):
            if w is not None:
                processed_chart_crit_names.append(temp_crit_names[i])
                processed_chart_crit_weights.append(float(w))

    processed_chart_alt_names = []
    processed_chart_alt_scores = []
    if alternatives_from_db:
        temp_alt_names_for_matching_scores = []
        for alt in alternatives_from_db:
            score = alt.get('overall_score')
            if score is not None:
                temp_alt_names_for_matching_scores.append(alt.get('alternative_name', 'N/A'))
                processed_chart_alt_scores.append(float(score))
        processed_chart_alt_names = temp_alt_names_for_matching_scores
        # Đảm bảo độ dài khớp nếu có phương án không có điểm
        if len(processed_chart_alt_scores) != len(processed_chart_alt_names) and processed_chart_alt_scores:
            final_alt_names = []
            score_idx = 0
            for alt_item_check in alternatives_from_db:
                if alt_item_check.get('overall_score') is not None:
                    if score_idx < len(processed_chart_alt_scores):
                         final_alt_names.append(alt_item_check.get('alternative_name', 'N/A'))
                    score_idx +=1
            processed_chart_alt_names = final_alt_names


    results_for_display = {
        'decision_info': decision_info,
        'criteria_names': [c.get('criterion_name', 'N/A') for c in criteria_from_db] if criteria_from_db else [],
        'criteria_weights': [c.get('weight') for c in criteria_from_db] if criteria_from_db else [],
        'cr': decision_info.get('criteria_cr'), # Đây là key đúng
        'comparison_matrix': crit_comp_matrix_list_from_db,
        'ranked_alternatives': alternatives_from_db,
        'num_alternatives_processed': len(alternatives_from_db),
        'chart_crit_names': processed_chart_crit_names,
        'chart_crit_weights': processed_chart_crit_weights,
        'chart_alt_names': processed_chart_alt_names,
        'chart_alt_scores': processed_chart_alt_scores
    }
    return render_template('decision_results.html', results=results_for_display)

@app.route('/history/delete/<int:decision_id>', methods=['POST'])
def delete_history_entry(decision_id):
    if db_manager.delete_decision_by_id(decision_id): flash(f"Đã xóa quyết định ID {decision_id} khỏi lịch sử.", "success")
    else: flash(f"Lỗi khi xóa quyết định ID {decision_id}.", "error")
    return redirect(url_for('history'))

# File: app.py
# ... (các import và code khác) ...

def _generate_excel_workbook(ahp_data_dict):
    """Hàm helper để tạo workbook Excel từ dữ liệu AHP."""
    wb = Workbook()
    # ... (phần tạo sheet và ghi dữ liệu ws1, ws2 giữ nguyên như code hoàn chỉnh lần trước)
    ws1 = wb.active; ws1.title = "TieuChi_AHP"
    # ... (ghi dữ liệu cho sheet Tiêu chí) ...
    # ... (ghi dữ liệu cho sheet Phương án) ...

    # Điều chỉnh độ rộng cột (ĐOẠN CẦN SỬA)
    for sheet_name_export in wb.sheetnames:
        ws_export = wb[sheet_name_export]
        for col_idx_export in range(1, ws_export.max_column + 1):
            column_letter_export = get_column_letter(col_idx_export)
            max_length_export = 0
            for row_idx_export in range(1, ws_export.max_row + 1):
                cell_export = ws_export.cell(row=row_idx_export, column=col_idx_export)
                if cell_export.value is not None:
                    try:
                        # Cố gắng lấy độ dài của giá trị ô sau khi chuyển thành chuỗi
                        cell_value_as_string = str(cell_export.value)
                        if len(cell_value_as_string) > max_length_export:
                            max_length_export = len(cell_value_as_string)
                    except Exception: # THÊM KHỐI EXCEPT ĐẦY ĐỦ
                        pass # Bỏ qua nếu không thể xử lý giá trị ô
            
            # Đặt độ rộng cột
            adjusted_width_export = (max_length_export + 2) if max_length_export > 0 else 12
            ws_export.column_dimensions[column_letter_export].width = adjusted_width_export
    return wb

# ... (các route và hàm khác giữ nguyên) ...

# if __name__ == '__main__':
#    # ... (code khởi động) ...

@app.route('/export_excel_current', methods=['POST'])
def export_excel_current():
    try:
        criteria_data = _calculate_criteria_logic(request.form)
        full_ahp_data = _calculate_full_ahp_logic(request.form, criteria_data)
        wb = _generate_excel_workbook(full_ahp_data)
        excel_io=io.BytesIO(); wb.save(excel_io); excel_io.seek(0); ts=datetime.now().strftime("%Y%m%d_%H%M%S"); fn=f"AHP_KetQuaHienTai_{ts}.xlsx"
        return send_file(excel_io,as_attachment=True,download_name=fn,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except ValueError as e: flash(f"Lỗi dữ liệu khi xuất Excel: {e}","error"); return redirect(url_for('index'))
    except Exception as e_gen: flash(f"Lỗi không xác định khi xuất Excel: {e_gen}","error"); return redirect(url_for('index'))

@app.route('/export_excel_history/<int:decision_id>', methods=['GET'])
def export_excel_history(decision_id):
    # ... (Code của hàm này giữ nguyên như phiên bản đầy đủ trước đó, chỉ cần đảm bảo
    #      phần tạo ahp_data_from_history và gọi _generate_excel_workbook là đúng) ...
    decision_info = db_manager.get_decision_by_id(decision_id);
    if not decision_info: flash(f"Không tìm thấy QĐ ID {decision_id}","error"); return redirect(url_for('history'))
    criteria_db = db_manager.get_criteria_for_decision(decision_id); alternatives_db = db_manager.get_alternatives_with_scores_for_decision(decision_id)
    crit_comp_matrix_list=None
    if decision_info.get('criteria_comparison_json'):
        try: crit_comp_matrix_list=json.loads(decision_info['criteria_comparison_json'])
        except (json.JSONDecodeError, TypeError): flash("Lỗi đọc ma trận JSON.","warning")
    
    # Chuẩn bị ahp_data_from_history một cách cẩn thận hơn
    ahp_data_from_history = {
        'criteria_names': [c.get('criterion_name', 'N/A') for c in criteria_db] if criteria_db else [],
        'comparison_matrix': crit_comp_matrix_list if crit_comp_matrix_list else [],
        'weights': [c.get('weight') for c in criteria_db] if criteria_db else [],
        'lambda_max': None, 'ci': None, # Không lưu các giá trị này trong DB hiện tại
        'cr': decision_info.get('criteria_cr'),
        'is_consistent': decision_info.get('criteria_cr') <= 0.1 if decision_info.get('criteria_cr') is not None else None,
        'normalized_matrix_by_col': [], # Không lưu chi tiết này
        'ranked_alternatives': [],
        'criteria_types': [] # Không lưu chi tiết này trong DB hiện tại
    }
    if alternatives_db:
        for alt_db_item in alternatives_db:
            # Sắp xếp lại scores_detail theo thứ tự của criteria_names trong ahp_data_from_history
            raw_scores_ordered = [None] * len(ahp_data_from_history['criteria_names'])
            normalized_scores_ordered = [None] * len(ahp_data_from_history['criteria_names'])
            if alt_db_item.get('scores_detail'):
                for score_item in alt_db_item['scores_detail']:
                    try:
                        idx = ahp_data_from_history['criteria_names'].index(score_item['criterion_name'])
                        raw_scores_ordered[idx] = score_item.get('raw_score')
                        normalized_scores_ordered[idx] = score_item.get('normalized_score')
                    except (ValueError, IndexError): # Tên tiêu chí không khớp hoặc index lỗi
                        pass 
            ahp_data_from_history['ranked_alternatives'].append({
                'rank': alt_db_item.get('rank_value'), 'name': alt_db_item.get('alternative_name'),
                'raw_scores': raw_scores_ordered, 'normalized_scores': normalized_scores_ordered,
                'overall_score': alt_db_item.get('overall_score')})
    try:
        wb = _generate_excel_workbook(ahp_data_from_history)
        excel_io=io.BytesIO(); wb.save(excel_io); excel_io.seek(0); ts=datetime.now().strftime("%Y%m%d_%H%M%S"); 
        fn_name_safe = "".join(c if c.isalnum() else "_" for c in decision_info.get('decision_name','LichSu'))
        fn=f"AHP_LichSu_{fn_name_safe}_{decision_id}_{ts}.xlsx"
        return send_file(excel_io,as_attachment=True,download_name=fn,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e: flash(f"Lỗi xuất Excel từ lịch sử: {str(e)}","error"); return redirect(url_for('history'))


if __name__ == '__main__':
    print("Chào mừng bạn đến với Ứng dụng Hỗ trợ Ra Quyết định AHP Mua Đồ Điện Tử!")
    test_conn = db_manager.create_connection()
    if test_conn: print("Kết nối CSDL (XAMPP với PyMySQL) thành công."); test_conn.close()
    else: print("LỖI NGHIÊM TRỌNG: Không thể kết nối đến cơ sở dữ liệu.")
    if test_conn: app.run(debug=True, host='0.0.0.0', port=5000)